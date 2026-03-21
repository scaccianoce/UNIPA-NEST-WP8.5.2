"""Main application entry point for geometry-driven EnergyPlus simulation, early-stage LCA evaluation, and discrete Pareto optimization on building clusters."""
from __future__ import annotations

import argparse
import csv
import sys
import datetime as dt
import itertools
import platform
import queue
import re
import subprocess
import threading
from collections import defaultdict
from dataclasses import dataclass, field, replace
from pathlib import Path
from typing import Any, Callable, Mapping, Optional


COORD_OFFSET_X = 2373860
COORD_OFFSET_Y = 4218530
ADJACENCY_BUFFER = 0.1
EDGE_SAMPLE_OFFSET = 0.2
DEFAULT_MAC_ENERGYPLUS_DIR = Path('/Applications/EnergyPlus-24-2-0')
DEFAULT_WINDOWS_ENERGYPLUS_DIR = Path(r'C:/EnergyPlus-23-2-0')
DEFAULT_EXCEL_NAME = 'Tipologie_edilizie_misure.xlsx'
DEFAULT_OPTIMIZATION_TEMPLATE = 'Template_input_ottimizzazione.xlsx'
DEFAULT_TEMPLATE_TYPOLOGY_SHEET = 'Tipologie_Intervento'
TEMPLATE_BASE_SHEET = 'Tipologie_Base'
TEMPLATE_INTERVENTION_SHEET = 'Tipologie_Intervento'
TEMPLATE_VARIABLES_SHEET = 'Variabili_Ottimizzazione'
TEMPLATE_MAPPING_SHEET = 'Mapping_EnergyPlus_LCA'
TEMPLATE_LCA_SHEET = 'Early_LCA_ClimateChange'
TEMPLATE_PARAMETERS_SHEET = 'Parametri_Analisi'
DEFAULT_IDF_TEMPLATE = 'filebase.idf'
DEFAULT_SHAPEFILE = 'shapefile/Cuba-Calatafimi_District.shp'
DEFAULT_WEATHER_DIR = 'ITA_Palermo'
DEFAULT_START_INDEX = 0
LOG_PREFIX = '[LCADUBS]'
GJ_TO_MWH = 1 / 3.6
INTERVENTION_COLUMNS = {
    'nome della costruzione del muro esterno': ('wall', 'pareti'),
    'nome della costruzione del tetto': ('roof', 'copertura'),
    'nome della costruzione della finestra': ('window', 'vetri'),
    'nome della costruzione del Telaio': ('frame', 'telai'),
}
REPORT_TABLE_METRICS = (
    ('Area', 41, 2),
    ('Natural Gas Heating', 49, 3),
    ('Electric Cooling', 50, 2),
    ('Electric Fan', 55, 2),
    ('Electric Pump', 56, 2),
)
REPORT_TIMESERIES_METRICS = (
    'Fan Coil Heating Energy',
    'Fan Coil Total Cooling Energy',
    'Fan Coil Fan Electricity Energy',
    'Sensible Heating Energy',
    'Sensible Cooling Energy',
)


@dataclass
class Point:
    """Simple 3D point used to build EnergyPlus coordinates."""
    x: float
    y: float
    z: float = 0.0

    def move(self, dx: float, dy: float, dz: float) -> None:
        """Translate the point in place by the provided Cartesian offsets."""
        self.x += dx
        self.y += dy
        self.z += dz

    def distance(self, other: 'Point') -> float:
        """Return the Euclidean distance between this point and another point."""
        dx = self.x - other.x
        dy = self.y - other.y
        dz = self.z - other.z
        return (dx ** 2 + dy ** 2 + dz ** 2) ** 0.5

    def nearest(self, p1: 'Point', p2: 'Point') -> 'Point':
        """Return whichever of the two candidate points is closer to the current point."""
        nearer = p1 if p1.distance(self) < p2.distance(self) else p2
        return Point(nearer.x, nearer.y, nearer.z)

    def getX(self) -> float:
        """Expose the x coordinate using the legacy accessor style used elsewhere in the file."""
        return self.x

    def getY(self) -> float:
        """Expose the y coordinate using the legacy accessor style used elsewhere in the file."""
        return self.y

    def getZ(self) -> float:
        """Expose the z coordinate using the legacy accessor style used elsewhere in the file."""
        return self.z

    def __str__(self) -> str:
        """Format the point as the comma-separated coordinate string expected by EnergyPlus."""
        return f'{self.x:.4f},{self.y:.4f},{self.z:.4f}'


@dataclass
class SimulationConfig:
    """Runtime options for input files, geometry thresholds and EnergyPlus execution."""
    excel_path: Path
    shapefile_path: Path
    idf_template_path: Path
    weather_files: list[Path]
    energyplus_dir: Path
    template_path: Optional[Path] = None
    typology_sheet: str = DEFAULT_TEMPLATE_TYPOLOGY_SHEET
    optimize: bool = False
    start_index: int = DEFAULT_START_INDEX
    end_index: Optional[int] = None
    output_dir: Optional[Path] = None
    min_wall_width: float = 2.0
    window_top_gap: float = 0.1
    standard_window_width: float = 1.2
    min_vertex_distance: float = 0.1
    coord_offset_x: float = COORD_OFFSET_X
    coord_offset_y: float = COORD_OFFSET_Y
    weather_profiles: tuple[str, ...] = ('Monthly', 'Hourly')
    generate_reports: bool = True
    base_dir: Path = field(default_factory=lambda: Path.cwd())


@dataclass
class BuildingTypology:
    """Construction and window properties loaded from the typology spreadsheet."""
    wall_construction: str
    ground_floor_construction: str
    roof_construction: str
    floor_construction: str
    ceiling_construction: str
    window_construction: str
    frame_construction: str
    schedule_name: str
    window_to_wall_ratio: float
    floor_height: float
    sill_height: float
    window_height: float


@dataclass(frozen=True)
class ScenarioDefinition:
    """Describe one discrete optimization scenario generated from baseline and intervention typology sheets."""
    id: str
    label: str
    selected_columns: tuple[str, ...]
    typology_data: Any


@dataclass
class BuildingGeometrySummary:
    """Store the simplified geometric quantities used by the early LCA engine."""
    footprint_area: float
    roof_area: float
    opaque_wall_area: float
    window_area: float
    frame_area: float


def default_energyplus_dir() -> Path:
    """Return the default EnergyPlus installation directory for the current operating system."""
    return DEFAULT_MAC_ENERGYPLUS_DIR if platform.system().upper() == 'DARWIN' else DEFAULT_WINDOWS_ENERGYPLUS_DIR


def discover_weather_files(weather_dir: Path) -> list[Path]:
    """Scan a directory and return all EPW weather files sorted by file name."""
    if not weather_dir.exists():
        return []
    return sorted(path for path in weather_dir.glob('*.epw') if path.is_file())


def default_template_path(base_dir: Path) -> Optional[Path]:
    """Return the default optimization template path if the workbook is present in the project folder."""
    candidate = base_dir / DEFAULT_OPTIMIZATION_TEMPLATE
    return candidate if candidate.exists() else None


def build_default_config(base_dir: Path) -> SimulationConfig:
    """Assemble the default configuration starting from the current working directory."""
    return SimulationConfig(
        excel_path=base_dir / DEFAULT_EXCEL_NAME,
        shapefile_path=base_dir / DEFAULT_SHAPEFILE,
        idf_template_path=base_dir / DEFAULT_IDF_TEMPLATE,
        weather_files=discover_weather_files(base_dir / DEFAULT_WEATHER_DIR),
        energyplus_dir=default_energyplus_dir(),
        template_path=default_template_path(base_dir),
        typology_sheet=DEFAULT_TEMPLATE_TYPOLOGY_SHEET,
        base_dir=base_dir,
    )


def log_message(message: str, callback: Optional[Callable[[str], None]] = None) -> None:
    """Send a log line either to a callback sink or to standard output."""
    line = f'{LOG_PREFIX} {message}'
    (callback or print)(line)


def sort_points_counterclockwise(coords: list[Point]) -> list[Point]:
    """Normalize polygon vertex order so that downstream surface generation is geometrically consistent."""
    area = 0.0
    for index in range(len(coords)):
        x1 = coords[index].getX()
        y1 = coords[index].getY()
        x2 = coords[(index + 1) % len(coords)].getX()
        y2 = coords[(index + 1) % len(coords)].getY()
        area += (x2 - x1) * (y2 + y1)
    return coords[::-1] if area > 0 else coords[:]


def require_runtime_dependencies():
    """Import optional runtime dependencies only when execution really needs them."""
    try:
        import pandas as pd
        import openpyxl  # noqa: F401
        import shapefile
        from shapely.geometry import LineString, Polygon
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "Dipendenze mancanti. Installa almeno: pandas, openpyxl, pyshp e shapely."
        ) from exc
    return pd, shapefile, LineString, Polygon


def require_report_dependencies():
    """Import optional reporting dependencies only when final report generation is requested."""
    try:
        import ezdxf
        from ezdxf.addons import Importer
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "Per i report finali serve anche ezdxf. Installa almeno: ezdxf."
        ) from exc
    return ezdxf, Importer


def field_values(reader, field_name: str) -> list:
    """Extract all values of a named attribute field from a shapefile reader."""
    field_names = [field[0] for field in reader.fields[1:]]
    if field_name not in field_names:
        raise KeyError(f'Campo shapefile non trovato: {field_name}')
    field_index = field_names.index(field_name)
    return [record[field_index] for record in reader.records()]


def translate_points(points: list[Point], config: SimulationConfig) -> list[tuple[float, float]]:
    """Shift geometric points back to the original coordinate system expected by Shapely operations."""
    return [(point.x + config.coord_offset_x, point.y + config.coord_offset_y) for point in points]


def point_on_segment(start: Point, end: Point, distance: float) -> Point:
    """Interpolate a point located at a given distance from the start of a segment."""
    length = start.distance(end)
    if length == 0:
        return Point(start.x, start.y, 0)
    ratio = distance / length
    return Point(start.x + (end.x - start.x) * ratio, start.y + (end.y - start.y) * ratio, 0)


def sampled_segment(start: Point, end: Point, inset: float) -> list[Point]:
    """Create a reduced set of sample points along a segment for adjacency detection."""
    length = start.distance(end)
    if length <= 1:
        return [start, end]
    return [point_on_segment(start, end, inset), point_on_segment(start, end, length - inset)]


def window_layout(wall_length: float, window_ratio_linear: float, standard_window_width: float) -> tuple[float, list[float]]:
    """Compute window width and offsets along a wall from the linearized window-to-wall ratio."""
    dim_linear_windows = wall_length * window_ratio_linear
    dim_linear_free_wall = wall_length - dim_linear_windows
    if dim_linear_free_wall > 0:
        window_count = max(1, round(dim_linear_windows / standard_window_width))
        window_width = dim_linear_windows / window_count
        spacing = dim_linear_free_wall / window_count
        offsets = [spacing / 2 + (window_width + spacing) * idx for idx in range(window_count)]
    else:
        window_width = dim_linear_windows
        offsets = [0.0]
    return window_width, offsets


def load_typology(typology_data: Any, building_type: str) -> BuildingTypology:
    """Read one typology row and convert it into a strongly structured BuildingTypology object."""
    properties = typology_data[typology_data['Tipologia'] == building_type]
    if properties.empty:
        raise ValueError(f'Tipologia edilizia non trovata nel file Excel: {building_type}')
    row = properties.iloc[0]
    return BuildingTypology(
        wall_construction=row['nome della costruzione del muro esterno'],
        ground_floor_construction=row['nome della costruzione del solaio controterra'],
        roof_construction=row['nome della costruzione del tetto'],
        floor_construction=row['nome della costruzione del solaio interpiano'],
        ceiling_construction=row['nome della costruzione del soffitto'],
        window_construction=row['nome della costruzione della finestra'],
        frame_construction=row['nome della costruzione del Telaio'],
        schedule_name=row['Schedule impianto'],
        window_to_wall_ratio=float(row['Perc_WWR']),
        floor_height=float(row['altezza interpiano [m]']),
        sill_height=float(row['altezza davanzale']),
        window_height=float(row['Altezza finestra']),
    )


def typology_source_label(config: SimulationConfig) -> str:
    """Return a human-readable description of the typology source used by the current run."""
    if config.template_path is not None:
        return f'{config.template_path} [foglio: {config.typology_sheet}]'
    return str(config.excel_path)


def load_typology_data(config: SimulationConfig, pd) -> Any:
    """Load typology data either from the legacy Excel file or from the selected template sheet."""
    if config.template_path is None:
        return pd.read_excel(config.excel_path, engine='openpyxl')
    try:
        return pd.read_excel(config.template_path, sheet_name=config.typology_sheet, engine='openpyxl')
    except ValueError as exc:
        raise ValueError(
            f'Foglio tipologie non trovato nel template {config.template_path}: {config.typology_sheet}'
        ) from exc


def require_optimization_template(config: SimulationConfig) -> Path:
    """Ensure that optimization mode has a valid external template workbook to work with."""
    if config.template_path is None:
        raise ValueError(
            'L ottimizzazione richiede un template Excel esterno. Seleziona Template_input_ottimizzazione.xlsx.'
        )
    return config.template_path


def load_template_sheet(config: SimulationConfig, pd, sheet_name: str) -> Any:
    """Load a named sheet from the optimization template workbook."""
    template_path = require_optimization_template(config)
    try:
        return pd.read_excel(template_path, sheet_name=sheet_name, engine='openpyxl')
    except ValueError as exc:
        raise ValueError(f'Foglio "{sheet_name}" non trovato nel template {template_path}.') from exc


def normalize_text(value) -> str:
    """Normalize text for case-insensitive and whitespace-tolerant comparisons."""
    return str(value).strip().lower() if value is not None else ''


def truthy(value) -> bool:
    """Interpret spreadsheet-style yes/no values as booleans."""
    return normalize_text(value) in {'1', 'true', 'si', 'sì', 'yes', 'y'}


def slugify_label(value: str) -> str:
    """Convert a free-text label into a stable slug suitable for file-system-safe scenario identifiers."""
    slug = re.sub(r'[^a-z0-9]+', '-', value.lower()).strip('-')
    return slug or 'scenario'


def parameter_dict(parameters_df) -> dict[str, str]:
    """Convert the parameter worksheet into a key-value dictionary."""
    if 'parametro' not in parameters_df.columns or 'valore' not in parameters_df.columns:
        return {}
    return {
        str(row['parametro']).strip(): '' if row['valore'] is None else str(row['valore']).strip()
        for _, row in parameters_df.iterrows()
        if row.get('parametro') is not None
    }


def parameter_float(parameters: Mapping[str, object], key: str, default: float) -> float:
    """Read one parameter value and convert it to float while preserving a safe fallback."""
    value = parameters.get(key)
    if value in (None, ''):
        return default
    try:
        return float(str(value).replace(',', '.'))
    except ValueError:
        return default


def parameter_text(parameters: Mapping[str, object], key: str, default: str) -> str:
    """Read one parameter value and convert it to text while preserving a safe fallback."""
    value = parameters.get(key)
    return value if value else default


def active_intervention_columns(base_df, intervention_df, variables_df) -> list[str]:
    """Detect which typology columns are both active in the template and actually changed by the intervention sheet."""
    active_columns = {
        str(row['campo_tipologia_coinvolto']).strip()
        for _, row in variables_df.iterrows()
        if truthy(row.get('attiva')) and row.get('campo_tipologia_coinvolto') is not None
    }
    available_columns: list[str] = []
    for column_name in INTERVENTION_COLUMNS:
        if column_name not in base_df.columns or column_name not in intervention_df.columns:
            continue
        if active_columns and column_name not in active_columns:
            continue
        base_values = base_df[column_name].fillna('').astype(str)
        intervention_values = intervention_df[column_name].fillna('').astype(str)
        if not base_values.equals(intervention_values):
            available_columns.append(column_name)
    return available_columns


def scenario_label(selected_columns: tuple[str, ...]) -> str:
    """Build the user-facing label associated with one intervention combination."""
    if not selected_columns:
        return 'baseline'
    return '+'.join(INTERVENTION_COLUMNS[column][1] for column in selected_columns)


def build_scenarios(base_df, intervention_df, variables_df) -> list[ScenarioDefinition]:
    """Generate the full discrete scenario space from the active intervention columns."""
    if 'Tipologia' not in base_df.columns or 'Tipologia' not in intervention_df.columns:
        raise ValueError('I fogli tipologici devono contenere la colonna "Tipologia".')
    base_indexed = base_df.copy().set_index('Tipologia', drop=False)
    intervention_indexed = intervention_df.copy().set_index('Tipologia', drop=False)
    if set(base_indexed.index) != set(intervention_indexed.index):
        raise ValueError('I fogli Tipologie_Base e Tipologie_Intervento devono avere le stesse tipologie.')

    # The discrete optimization space is built by toggling only the typology
    # columns that are both active in the template and effectively different.
    columns = active_intervention_columns(base_indexed, intervention_indexed, variables_df)
    scenarios: list[ScenarioDefinition] = []
    for size in range(len(columns) + 1):
        for selected in itertools.combinations(columns, size):
            scenario_df = base_indexed.copy()
            for column_name in selected:
                scenario_df.loc[:, column_name] = intervention_indexed.loc[scenario_df.index, column_name]
            label = scenario_label(selected)
            scenarios.append(
                ScenarioDefinition(
                    id=f'{len(scenarios):02d}_{slugify_label(label)}',
                    label=label,
                    selected_columns=selected,
                    typology_data=scenario_df.reset_index(drop=True),
                )
            )
    return scenarios


def polygon_area(points: list[Point], vertex_count: int) -> float:
    """Compute polygon area with the shoelace formula on already cleaned vertices."""
    area = 0.0
    for index in range(vertex_count):
        area += (
            points[index].x * points[index + 1].y
            - points[index + 1].x * points[index].y
        )
    return abs(area) / 2


def building_geometry_summary(
    building_index: int,
    shape,
    shapes: list,
    floor_count: int,
    typology: BuildingTypology,
    config: SimulationConfig,
) -> BuildingGeometrySummary:
    """Compute the simplified wall, roof, window, and frame quantities used by the LCA engine."""
    vertices, vertex_count = clean_vertices(shape.points, config)
    adjacent_flags = adjacent_walls(building_index, vertices, shapes, config)
    edge_lengths = [vertices[idx].distance(vertices[idx + 1]) for idx in range(vertex_count)]
    exposed_perimeter = sum(length for idx, length in enumerate(edge_lengths) if not adjacent_flags[idx])
    perimeter = exposed_perimeter if exposed_perimeter > 0 else sum(edge_lengths)
    gross_wall_area = perimeter * typology.floor_height * floor_count
    window_area = gross_wall_area * typology.window_to_wall_ratio
    opaque_wall_area = max(gross_wall_area - window_area, 0.0)
    footprint_area = polygon_area(vertices, vertex_count)
    return BuildingGeometrySummary(
        footprint_area=footprint_area,
        roof_area=footprint_area,
        opaque_wall_area=opaque_wall_area,
        window_area=window_area,
        frame_area=window_area,
    )


def variable_defaults_by_column(variables_df) -> dict[str, dict[str, float | str]]:
    """Group optimization-variable defaults by the typology column they control."""
    defaults: dict[str, dict[str, float | str]] = {}
    for _, row in variables_df.iterrows():
        column_name = row.get('campo_tipologia_coinvolto')
        if column_name is None:
            continue
        defaults[str(column_name).strip()] = {
            'id_variabile': '' if row.get('id_variabile') is None else str(row.get('id_variabile')).strip(),
            'valore_min': row.get('valore_min'),
            'valore_max': row.get('valore_max'),
            'step': row.get('step'),
        }
    return defaults


def parse_quantity_expression(expression: str, context: dict[str, float]) -> float:
    # Keep the evaluator explicit and closed: only template expressions that are
    # documented and supported by the tool are accepted here.
    """Evaluate the restricted set of quantity expressions supported by the early LCA template."""
    normalized = normalize_text(expression)
    if normalized == 'area * spessore * densita':
        return context['area'] * context['spessore'] * context['densita']
    if normalized == 'area * spessore_rimosso':
        return context['area'] * context['spessore_rimosso']
    if normalized == 'area_vetro':
        return context['area_vetro']
    if normalized == 'area_telaio o area_finestra':
        return context['area_telaio']
    if normalized == 'area_vetro * spessore_vetro * densita':
        return context['area_vetro'] * context['spessore_vetro'] * context['densita']
    if normalized == 'area / cover':
        return context['area'] / context['cover'] if context['cover'] else 0.0
    if normalized == 'energia_termica_annua':
        return context['energia_termica_annua']
    if normalized == 'energia_elettrica_annua':
        return context['energia_elettrica_annua']
    raise ValueError(f'Espressione di quantita non supportata: {expression}')


def parse_optional_float(value, default: float) -> float:
    """Convert an optional numeric field to float while returning a default on invalid input."""
    if value is None or value == '':
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def aggregate_energy_results(output_dir: Path, weather_files: list[Path]) -> dict[str, float]:
    """Aggregate EnergyPlus summary metrics over all buildings and all weather files for one scenario."""
    building_names = sorted((path.stem for path in output_dir.glob('Edificio*.idf')), key=building_sort_key)
    if not building_names:
        raise FileNotFoundError(f'Nessun file IDF trovato in {output_dir} per aggregare i risultati.')

    weather_totals: list[dict[str, float]] = []
    for weather_index, _ in enumerate(weather_files):
        totals = defaultdict(float)
        for building_name in building_names:
            weather_output_dir = output_dir / f'{building_name}W{weather_index}'
            table_values = extract_table_summary(weather_output_dir)
            for metric_name, value in table_values.items():
                totals[metric_name] += value
        weather_totals.append(dict(totals))

    aggregation = 'mean'
    final = defaultdict(float)
    for totals in weather_totals:
        for metric_name, value in totals.items():
            final[metric_name] += value
    if aggregation == 'mean' and weather_totals:
        for metric_name in list(final.keys()):
            final[metric_name] /= len(weather_totals)

    electricity_gj = final['Electric Cooling'] + final['Electric Fan'] + final['Electric Pump']
    gas_gj = final['Natural Gas Heating']
    return {
        'area_total_m2': final['Area'],
        'natural_gas_gj': gas_gj,
        'electricity_gj': electricity_gj,
        'natural_gas_mwh': gas_gj * GJ_TO_MWH,
        'electricity_mwh': electricity_gj * GJ_TO_MWH,
        'energy_total_gj': gas_gj + electricity_gj,
        'energy_total_mwh': (gas_gj + electricity_gj) * GJ_TO_MWH,
    }


def lca_factor_table(lca_df) -> dict[tuple[str, str, str], dict[str, float | str]]:
    """Index early-database LCA factors by material, category, and life-cycle phase."""
    factors: dict[tuple[str, str, str], dict[str, float | str]] = {}
    for _, row in lca_df.iterrows():
        material = normalize_text(row.get('Material'))
        category = normalize_text(row.get('Category'))
        phase = normalize_text(row.get('Phase'))
        if not material or not category or not phase:
            continue
        factors[(material, category, phase)] = {
            'climate_change': parse_optional_float(row.get('Climate change'), 0.0),
            'density': parse_optional_float(row.get('Average density'), 0.0),
            'cover': parse_optional_float(row.get('Cover'), 0.0),
        }
    return factors


def mapping_rows_for(mapping_df, energyplus_category: str, energyplus_name: str, phase: str) -> list[dict[str, object]]:
    """Return all template mapping rows matching one EnergyPlus category, name, and LCA phase."""
    category_key = normalize_text(energyplus_category)
    name_key = normalize_text(energyplus_name)
    phase_key = normalize_text(phase)
    rows: list[dict[str, object]] = []
    for _, row in mapping_df.iterrows():
        if (
            normalize_text(row.get('energyplus_categoria')) == category_key
            and normalize_text(row.get('energyplus_nome')) == name_key
            and normalize_text(row.get('fase_lca')) == phase_key
        ):
            rows.append(row.to_dict())
    return rows


def lca_amount_for_mapping(
    mapping_row: dict[str, object],
    geometry: BuildingGeometrySummary,
    defaults: dict[str, dict[str, float | str]],
    factors: dict[tuple[str, str, str], dict[str, float | str]],
    selected_column: str,
    energy_context: dict[str, float],
) -> float:
    # Each LCA row combines: geometry-driven quantity, optional overrides from
    # the mapping sheet, and the early-database impact factor for the phase.
    """Calculate the climate-change contribution associated with one mapping row and one scenario context."""
    material = normalize_text(mapping_row.get('materiale_lca'))
    category = normalize_text(mapping_row.get('categoria_lca'))
    phase = normalize_text(mapping_row.get('fase_lca'))
    factor_info = factors.get((material, category, phase))
    if factor_info is None:
        raise ValueError(
            f'Fattore LCA mancante per materiale={mapping_row.get("materiale_lca")}, '
            f'categoria={mapping_row.get("categoria_lca")}, fase={mapping_row.get("fase_lca")}.'
        )
    variable_info = defaults.get(selected_column, {})
    thickness = max(parse_optional_float(variable_info.get('valore_max'), 0.0), 0.0)
    intervention_key = INTERVENTION_COLUMNS[selected_column][0]
    if intervention_key == 'wall':
        area_value = geometry.opaque_wall_area
    elif intervention_key == 'roof':
        area_value = geometry.roof_area
    else:
        area_value = geometry.window_area
    context = {
        'area': area_value,
        'area_vetro': geometry.window_area,
        'area_telaio': geometry.frame_area,
        'spessore': thickness,
        'spessore_rimosso': parameter_float(energy_context, 'opaque_disposal_thickness_m', 0.0),
        'spessore_vetro': parameter_float(energy_context, 'glass_disposal_thickness_m', 0.024),
        'densita': parse_optional_float(mapping_row.get('densita_override'), 0.0)
        or float(factor_info['density']),
        'cover': parse_optional_float(mapping_row.get('cover_override'), 0.0)
        or float(factor_info['cover'])
        or 1.0,
        'energia_termica_annua': energy_context['energia_termica_annua'],
        'energia_elettrica_annua': energy_context['energia_elettrica_annua'],
    }
    quantity = parse_quantity_expression(str(mapping_row.get('base_quantita', '')), context)
    return quantity * float(factor_info['climate_change'])


def pareto_flags(rows: list[dict[str, float | str]]) -> list[bool]:
    """Mark which scenarios are non-dominated in the bi-objective energy-versus-LCA space."""
    flags: list[bool] = []
    for row in rows:
        dominated = False
        for other in rows:
            if other is row:
                continue
            if (
                float(other['energy_total_mwh']) <= float(row['energy_total_mwh'])
                and float(other['lca_total_kgco2eq']) <= float(row['lca_total_kgco2eq'])
                and (
                    float(other['energy_total_mwh']) < float(row['energy_total_mwh'])
                    or float(other['lca_total_kgco2eq']) < float(row['lca_total_kgco2eq'])
                )
            ):
                dominated = True
                break
        flags.append(not dominated)
    return flags


def scenario_intervention_rows(
    scenario: ScenarioDefinition,
    base_df: Any,
    scenario_df: Any,
) -> list[dict[str, str]]:
    """Expand one scenario into a row-level description of which typologies were modified and how."""
    base_indexed = base_df.set_index('Tipologia', drop=False)
    scenario_indexed = scenario_df.set_index('Tipologia', drop=False)
    rows: list[dict[str, str]] = []
    for typology in sorted(base_indexed.index):
        for column_name in scenario.selected_columns:
            base_value = base_indexed.loc[typology, column_name]
            scenario_value = scenario_indexed.loc[typology, column_name]
            if base_value == scenario_value:
                continue
            rows.append(
                {
                    'scenario_id': scenario.id,
                    'scenario_label': scenario.label,
                    'tipologia': str(typology),
                    'componente': INTERVENTION_COLUMNS[column_name][1],
                    'campo_tipologia': column_name,
                    'valore_base': str(base_value),
                    'valore_intervento': str(scenario_value),
                }
            )
    if not rows:
        rows.append(
            {
                'scenario_id': scenario.id,
                'scenario_label': scenario.label,
                'tipologia': 'tutte',
                'componente': 'baseline',
                'campo_tipologia': '',
                'valore_base': '',
                'valore_intervento': '',
            }
        )
    return rows


def add_baseline_deltas(scenario_rows: list[dict[str, float | str]]) -> tuple[list[dict[str, float | str]], dict[str, float]]:
    """Add energy-saving and LCA-reduction deltas with respect to the baseline scenario."""
    baseline_row = next((row for row in scenario_rows if str(row.get('scenario_label')) == 'baseline'), None)
    if baseline_row is None:
        raise ValueError('Scenario baseline non trovato nei risultati di ottimizzazione.')

    baseline_energy = float(baseline_row['energy_total_mwh'])
    baseline_lca = float(baseline_row['lca_total_kgco2eq'])
    baseline_summary = {
        'baseline_energy_mwh': baseline_energy,
        'baseline_lca_kgco2eq': baseline_lca,
    }
    for row in scenario_rows:
        energy = float(row['energy_total_mwh'])
        lca_total = float(row['lca_total_kgco2eq'])
        energy_saved = baseline_energy - energy
        lca_reduction = baseline_lca - lca_total
        row['energy_saved_mwh_vs_baseline'] = round(energy_saved, 6)
        row['energy_saved_pct_vs_baseline'] = round((energy_saved / baseline_energy * 100) if baseline_energy else 0.0, 6)
        row['lca_final_kgco2eq'] = round(lca_total, 6)
        row['lca_reduction_kgco2eq_vs_baseline'] = round(lca_reduction, 6)
        row['lca_reduction_pct_vs_baseline'] = round((lca_reduction / baseline_lca * 100) if baseline_lca else 0.0, 6)
    return scenario_rows, baseline_summary


def pareto_summary_rows(scenario_rows: list[dict[str, float | str]]) -> list[dict[str, float | str]]:
    """Extract the compact subset of result fields used for the Pareto summary sheet."""
    return [
        {
            'scenario_id': row['scenario_id'],
            'scenario_label': row['scenario_label'],
            'interventi': row['interventi'],
            'energy_total_mwh': row['energy_total_mwh'],
            'energy_saved_mwh_vs_baseline': row['energy_saved_mwh_vs_baseline'],
            'energy_saved_pct_vs_baseline': row['energy_saved_pct_vs_baseline'],
            'lca_final_kgco2eq': row['lca_final_kgco2eq'],
            'lca_reduction_kgco2eq_vs_baseline': row['lca_reduction_kgco2eq_vs_baseline'],
            'lca_reduction_pct_vs_baseline': row['lca_reduction_pct_vs_baseline'],
        }
        for row in scenario_rows
        if row.get('pareto') == 'SI'
    ]


def floor_levels(floor_count: int, floor_height: float) -> list[float]:
    """Return the elevation of each floor slab used when building vertical geometry."""
    return [level * floor_height for level in range(floor_count)]


def compute_window_parameters(
    vertices: list[Point],
    vertex_count: int,
    adjacent_flags: list[bool],
    typology: BuildingTypology,
    config: SimulationConfig,
    building_index: int,
) -> tuple[float, float, float]:
    """Adapt sill/window height and return the linear WWR used to distribute windows."""
    perimeter = sum(vertices[idx].distance(vertices[idx + 1]) for idx in range(vertex_count))
    total_window_area = perimeter * typology.floor_height * typology.window_to_wall_ratio
    available_wall_length = sum(
        vertices[idx].distance(vertices[idx + 1])
        for idx in range(vertex_count)
        if vertices[idx].distance(vertices[idx + 1]) > config.min_wall_width and not adjacent_flags[idx]
    )
    if available_wall_length <= 0:
        raise ValueError(f'Edificio {building_index}: nessuna parete disponibile per inserire finestre.')

    sill_height = typology.sill_height
    window_height = typology.window_height
    if available_wall_length * window_height < total_window_area:
        window_height = total_window_area / available_wall_length
        if window_height + config.window_top_gap + sill_height > typology.floor_height:
            sill_height = typology.floor_height - config.window_top_gap - window_height
            if sill_height < 0:
                sill_height = config.window_top_gap
                total_window_area = available_wall_length * (typology.floor_height - config.window_top_gap - sill_height)

    window_ratio_linear = total_window_area / window_height / available_wall_length
    return sill_height, window_height, window_ratio_linear


def build_zone_lines(floor_count: int) -> list[str]:
    """Generate the EnergyPlus zone declarations required for the current building."""
    lines = [f'Zone,Piano{floor},0,0,0,0,1,1,autocalculate,autocalculate,autocalculate;\n' for floor in range(floor_count)]
    zone_names = ''.join(f',Piano{floor}' for floor in range(floor_count))
    lines.append(f'ZoneList,listazone{zone_names};\n')
    return lines


def horizontal_surface_line(
    floor: int,
    surface_kind: str,
    construction: str,
    zone_name: str,
    outside_boundary: str,
    outside_boundary_object: str,
    sun_exposure: str,
    wind_exposure: str,
    vertices: list[Point],
    elevation: float,
    reverse_order: bool = False,
) -> str:
    """Create one EnergyPlus floor, roof, or ceiling surface definition."""
    ordered_vertices = reversed(vertices) if reverse_order else vertices
    coords = ''.join(f',{Point(vertex.x, vertex.y, elevation)}' for vertex in ordered_vertices)
    return (
        f'BuildingSurface:Detailed,{zone_name}:{surface_kind},{surface_kind},{construction},{zone_name},,'
        f'{outside_boundary},{outside_boundary_object},{sun_exposure},{wind_exposure},AutoCalculate,{len(vertices)}'
        f'{coords};\n'
    )


def build_floor_and_roof_lines(
    floor: int,
    vertex_count: int,
    vertices: list[Point],
    floor_levels: list[float],
    typology: BuildingTypology,
) -> list[str]:
    """Generate all horizontal envelope surfaces for one floor of the building."""
    zone_name = f'Piano{floor}'
    polygon = vertices[:vertex_count]
    floor_line = horizontal_surface_line(
        floor=floor,
        surface_kind='floor',
        construction=typology.ground_floor_construction if floor == 0 else typology.floor_construction,
        zone_name=zone_name,
        outside_boundary='Ground' if floor == 0 else 'Surface',
        outside_boundary_object='' if floor == 0 else f'Piano{floor - 1}:ceiling',
        sun_exposure='NoSun',
        wind_exposure='NoWind',
        vertices=polygon,
        elevation=floor_levels[floor],
        reverse_order=True,
    )
    roof_line = horizontal_surface_line(
        floor=floor,
        surface_kind='roof' if floor == len(floor_levels) - 1 else 'ceiling',
        construction=typology.roof_construction if floor == len(floor_levels) - 1 else typology.ceiling_construction,
        zone_name=zone_name,
        outside_boundary='Outdoors' if floor == len(floor_levels) - 1 else 'Surface',
        outside_boundary_object='' if floor == len(floor_levels) - 1 else f'Piano{floor + 1}:floor',
        sun_exposure='SunExposed' if floor == len(floor_levels) - 1 else 'NoSun',
        wind_exposure='WindExposed' if floor == len(floor_levels) - 1 else 'NoWind',
        vertices=polygon,
        elevation=floor_levels[floor] + typology.floor_height,
    )
    return [floor_line, roof_line]


def wall_surface_line(
    floor: int,
    side_index: int,
    vertices: list[Point],
    floor_levels: list[float],
    floor_height: float,
    wall_construction: str,
) -> str:
    """Create one EnergyPlus vertical wall surface definition."""
    v1 = Point(vertices[side_index].x, vertices[side_index].y, floor_levels[floor] + floor_height)
    v2 = Point(vertices[side_index].x, vertices[side_index].y, floor_levels[floor])
    v3 = Point(vertices[side_index + 1].x, vertices[side_index + 1].y, floor_levels[floor])
    v4 = Point(vertices[side_index + 1].x, vertices[side_index + 1].y, floor_levels[floor] + floor_height)
    wall_name = f'Muro{side_index}'
    return (
        f'BuildingSurface:Detailed,Piano{floor}:{wall_name},wall,{wall_construction},Piano{floor},,Outdoors,,'
        f'SunExposed,WindExposed,AutoCalculate,4,{v1},{v2},{v3},{v4};\n'
    )


def build_envelope_lines(
    vertex_count: int,
    vertices: list[Point],
    floor_count: int,
    floor_levels: list[float],
    typology: BuildingTypology,
    sill_height: float,
    window_height: float,
    window_ratio_linear: float,
    adjacent_flags: list[bool],
    config: SimulationConfig,
) -> list[str]:
    """Generate all opaque and transparent envelope surfaces for the building."""
    lines: list[str] = []
    for floor in range(floor_count):
        lines.extend(build_floor_and_roof_lines(floor, vertex_count, vertices, floor_levels, typology))
        for side_index in range(vertex_count):
            wall_name = f'Muro{side_index}'
            lines.append(
                wall_surface_line(
                    floor=floor,
                    side_index=side_index,
                    vertices=vertices,
                    floor_levels=floor_levels,
                    floor_height=typology.floor_height,
                    wall_construction=typology.wall_construction,
                )
            )
            lines.extend(
                build_window_lines(
                    floor_index=floor,
                    side_index=side_index,
                    wall_name=wall_name,
                    vertices=vertices,
                    floor_heights=floor_levels,
                    sill_height=sill_height,
                    window_height=window_height,
                    window_ratio_linear=window_ratio_linear,
                    wall_length_min=config.min_wall_width,
                    standard_window_width=config.standard_window_width,
                    adjacent_flags=adjacent_flags,
                    construction_window=typology.window_construction,
                    construction_frame=typology.frame_construction,
                )
            )
    return lines


def adjacent_walls(polygon_index: int, vertices: list[Point], shapes: list, config: SimulationConfig) -> list[bool]:
    """Detect which wall segments are adjacent to neighboring buildings and should be treated as non-exposed."""
    _, _, LineString, Polygon = require_runtime_dependencies()
    flags: list[bool] = []
    for side_index in range(len(vertices) - 1):
        sampled_points = sampled_segment(vertices[side_index], vertices[side_index + 1], EDGE_SAMPLE_OFFSET)
        dilated = LineString(translate_points(sampled_points, config)).buffer(ADJACENCY_BUFFER, cap_style='flat')
        flags.append(
            any(
                other_index != polygon_index and dilated.overlaps(Polygon(other_shape.points))
                for other_index, other_shape in enumerate(shapes)
            )
        )
    return flags


def unique_output_dir(base_dir: Path) -> Path:
    """Create a timestamped output directory without overwriting previous runs."""
    timestamp = dt.datetime.now().strftime('Simul_%d-%m_%H-%M')
    candidate = base_dir / timestamp
    counter = 1
    while candidate.exists():
        candidate = base_dir / f'{timestamp}_{counter}'
        counter += 1
    candidate.mkdir(parents=True, exist_ok=False)
    return candidate


def building_sort_key(name: str) -> tuple[int, str]:
    """Provide a stable numeric-aware sort key for building file names."""
    match = re.search(r'(\d+)$', name)
    if match is None:
        return sys.maxsize, name
    return int(match.group(1)), name


def weather_labels(weather_files: list[Path]) -> list[str]:
    """Create human-readable labels for all selected weather files."""
    labels: list[str] = []
    for index, weather_file in enumerate(weather_files):
        label = weather_file.stem.strip() or f'Weather{index}'
        if label in labels:
            label = f'{label}_{index}'
        labels.append(label)
    return labels


def float_or_zero(value: str) -> float:
    """Parse a numeric CSV field and fall back to zero on empty or invalid values."""
    try:
        return float(value.strip())
    except (TypeError, ValueError, AttributeError):
        return 0.0


def read_csv_rows(path: Path) -> list[list[str]]:
    """Read a CSV file as a raw list of rows."""
    if not path.exists() or path.stat().st_size == 0:
        return []
    with path.open('r', encoding='utf-8-sig', newline='') as handle:
        return list(csv.reader(handle))


def extract_table_summary(weather_output_dir: Path) -> dict[str, float]:
    """Extract key EnergyPlus summary-table metrics from eplustbl.csv."""
    rows = read_csv_rows(weather_output_dir / 'eplustbl.csv')
    values: dict[str, float] = {}
    for metric_name, row_index, column_index in REPORT_TABLE_METRICS:
        if row_index < len(rows) and column_index < len(rows[row_index]):
            values[metric_name] = float_or_zero(rows[row_index][column_index])
        else:
            values[metric_name] = 0.0
    return values


def extract_timeseries_summary(weather_output_dir: Path, pd) -> dict[str, float]:
    """Aggregate selected hourly or monthly time-series outputs from eplusout.csv."""
    csv_path = weather_output_dir / 'eplusout.csv'
    values = {metric_name: 0.0 for metric_name in REPORT_TIMESERIES_METRICS}
    if not csv_path.exists() or csv_path.stat().st_size == 0:
        return values

    data = pd.read_csv(csv_path)
    headers = data.columns.tolist()
    for metric_name in REPORT_TIMESERIES_METRICS:
        matching_headers = [header for header in headers if metric_name in header]
        values[metric_name] = sum(float(data[header].sum()) for header in matching_headers) / 1e9
    return values


def merge_dxf_document(source_doc, target_doc, importer_cls) -> None:
    """Import one DXF modelspace into the global merged DXF document."""
    importer = importer_cls(source_doc, target_doc)
    importer.import_modelspace()
    importer.finalize()


def write_summary_reports(output_dir: Path, weather_files: list[Path], callback: Optional[Callable[[str], None]] = None) -> list[Path]:
    """Create consolidated CSV, Excel, and DXF outputs for a standard simulation run."""
    pd, _, _, _ = require_runtime_dependencies()
    ezdxf, importer_cls = require_report_dependencies()

    building_names = sorted((path.stem for path in output_dir.glob('Edificio*.idf')), key=building_sort_key)
    if not building_names:
        raise FileNotFoundError(f'Nessun file IDF trovato in {output_dir} per generare i report.')

    labels = weather_labels(weather_files)
    table_rows: list[dict[str, float | str]] = []
    timeseries_rows: list[dict[str, float | str]] = []
    merged_doc = None

    for building_name in building_names:
        table_row: dict[str, float | str] = {'Edificio': building_name}
        timeseries_row: dict[str, float | str] = {'Edificio': building_name}
        for weather_index, weather_label in enumerate(labels):
            weather_output_dir = output_dir / f'{building_name}W{weather_index}'
            table_values = extract_table_summary(weather_output_dir)
            timeseries_values = extract_timeseries_summary(weather_output_dir, pd)
            for metric_name, value in table_values.items():
                table_row[f'{metric_name} [{weather_label}]'] = value
            for metric_name, value in timeseries_values.items():
                timeseries_row[f'{metric_name} [{weather_label}]'] = value

            if weather_index == 0:
                dxf_path = weather_output_dir / 'eplusout.dxf'
                if dxf_path.exists() and dxf_path.stat().st_size > 0:
                    if merged_doc is None:
                        merged_doc = ezdxf.readfile(dxf_path)
                    else:
                        merge_dxf_document(ezdxf.readfile(dxf_path), merged_doc, importer_cls)

        table_rows.append(table_row)
        timeseries_rows.append(timeseries_row)

    table_df = pd.DataFrame(table_rows)
    timeseries_df = pd.DataFrame(timeseries_rows)
    table_csv_path = output_dir / 'outputtbl.csv'
    workbook_path = output_dir / 'output.xlsx'
    table_df.to_csv(table_csv_path, index=False)
    with pd.ExcelWriter(workbook_path, engine='openpyxl') as writer:
        table_df.to_excel(writer, sheet_name='TableSummary', index=False)
        timeseries_df.to_excel(writer, sheet_name='TimeSeriesSummary', index=False)

    created_paths = [table_csv_path, workbook_path]
    if merged_doc is not None:
        merged_dxf_path = output_dir / 'merged.dxf'
        merged_doc.saveas(merged_dxf_path)
        created_paths.append(merged_dxf_path)

    log_message(
        'Report finali generati: ' + ', '.join(path.name for path in created_paths),
        callback,
    )
    return created_paths


def energyplus_paths(config: SimulationConfig) -> tuple[Path, Path]:
    """Resolve the EnergyPlus and ReadVarsESO executable paths from the configured installation directory."""
    is_windows = platform.system().upper().startswith('WIN')
    energyplus_name = 'energyplus.exe' if is_windows else 'energyplus'
    readvars_name = 'ReadVarsESO.exe' if is_windows else 'ReadVarsESO'
    energyplus_exe = config.energyplus_dir / energyplus_name
    readvars_exe = config.energyplus_dir / 'PostProcess' / readvars_name
    return energyplus_exe, readvars_exe


def validate_config(config: SimulationConfig) -> None:
    """Validate files, sheets, indices, and executable availability before launching a run."""
    require_runtime_dependencies()
    if config.generate_reports:
        require_report_dependencies()
    if config.optimize:
        require_optimization_template(config)
    required_paths = [config.shapefile_path, config.idf_template_path]
    if config.template_path is not None:
        required_paths.append(config.template_path)
    else:
        required_paths.append(config.excel_path)
    missing = [str(path) for path in required_paths if not path.exists()]
    if missing:
        raise FileNotFoundError('File richiesti mancanti:\n- ' + '\n- '.join(missing))
    if config.template_path is not None:
        from openpyxl import load_workbook

        workbook = load_workbook(config.template_path, data_only=True, read_only=True)
        try:
            if config.typology_sheet not in workbook.sheetnames:
                raise ValueError(
                    f'Foglio tipologie "{config.typology_sheet}" non trovato in {config.template_path}. '
                    f'Fogli disponibili: {", ".join(workbook.sheetnames)}'
                )
            if config.optimize:
                required_sheets = {
                    TEMPLATE_BASE_SHEET,
                    TEMPLATE_INTERVENTION_SHEET,
                    TEMPLATE_VARIABLES_SHEET,
                    TEMPLATE_MAPPING_SHEET,
                    TEMPLATE_LCA_SHEET,
                    TEMPLATE_PARAMETERS_SHEET,
                }
                missing_sheets = sorted(required_sheets.difference(workbook.sheetnames))
                if missing_sheets:
                    raise ValueError(
                        'Il template di ottimizzazione non contiene tutti i fogli richiesti:\n- '
                        + '\n- '.join(missing_sheets)
                    )
        finally:
            workbook.close()
    if not config.weather_files:
        raise FileNotFoundError('Nessun file meteo selezionato.')
    absent_weather = [str(path) for path in config.weather_files if not path.exists()]
    if absent_weather:
        raise FileNotFoundError('File meteo mancanti:\n- ' + '\n- '.join(absent_weather))
    if not config.energyplus_dir.exists():
        raise FileNotFoundError(f'Cartella EnergyPlus non trovata: {config.energyplus_dir}')
    energyplus_exe, readvars_exe = energyplus_paths(config)
    if not energyplus_exe.exists():
        raise FileNotFoundError(f'Eseguibile EnergyPlus non trovato: {energyplus_exe}')
    if not readvars_exe.exists():
        raise FileNotFoundError(f'Eseguibile ReadVarsESO non trovato: {readvars_exe}')
    if config.start_index < 0:
        raise ValueError('start_index deve essere >= 0.')
    if config.end_index is not None and config.end_index < config.start_index:
        raise ValueError('end_index deve essere >= start_index.')


def clean_vertices(shape_points: list[tuple[float, float]], config: SimulationConfig) -> tuple[list[Point], int]:
    """Remove degenerate or repeated vertices and close the cleaned polygon ring."""
    vertices = sort_points_counterclockwise(
        [Point(x - config.coord_offset_x, y - config.coord_offset_y, 0) for x, y in shape_points]
    )
    index = 0
    while index < len(vertices) - 1:
        if vertices[index].distance(vertices[index + 1]) < config.min_vertex_distance:
            del vertices[index]
            continue
        index += 1
    if len(vertices) > 1 and vertices[-1].distance(vertices[0]) < config.min_vertex_distance:
        del vertices[-1]
    if len(vertices) < 3:
        raise ValueError('Poligono con meno di 3 vertici utili dopo la pulizia.')
    vertex_count = len(vertices)
    vertices.append(vertices[0])
    return vertices, vertex_count


def build_window_lines(
    floor_index: int,
    side_index: int,
    wall_name: str,
    vertices: list[Point],
    floor_heights: list[float],
    sill_height: float,
    window_height: float,
    window_ratio_linear: float,
    wall_length_min: float,
    standard_window_width: float,
    adjacent_flags: list[bool],
    construction_window: str,
    construction_frame: str,
) -> list[str]:
    """Generate all EnergyPlus fenestration objects associated with one wall segment."""
    wall_start = vertices[side_index]
    wall_end = vertices[side_index + 1]
    wall_length = wall_start.distance(wall_end)
    if wall_length <= wall_length_min or adjacent_flags[side_index]:
        return []

    window_width, offsets = window_layout(wall_length, window_ratio_linear, standard_window_width)
    result: list[str] = []
    for count, distance_from_start in enumerate(offsets, start=1):
        start_edge = point_on_segment(wall_start, wall_end, distance_from_start)
        end_edge = point_on_segment(wall_start, wall_end, distance_from_start + window_width)

        v1 = Point(start_edge.x, start_edge.y, floor_heights[floor_index] + sill_height + window_height)
        v2 = Point(start_edge.x, start_edge.y, floor_heights[floor_index] + sill_height)
        v3 = Point(end_edge.x, end_edge.y, floor_heights[floor_index] + sill_height)
        v4 = Point(end_edge.x, end_edge.y, floor_heights[floor_index] + sill_height + window_height)

        result.append(
            'FenestrationSurface:Detailed,'
            f'Piano{floor_index}:{wall_name}:Fin{count},window,{construction_window},'
            f'Piano{floor_index}:{wall_name},,AutoCalculate,{construction_frame},1,4,'
            f'{v1},{v2},{v3},{v4};\n'
        )
    return result


def branch_list_line(system_name: str, branch_suffix: str, floor_count: int) -> str:
    """Create one EnergyPlus hydronic branch list line."""
    branches = ''.join(f',Piano{floor}:{branch_suffix}' for floor in range(floor_count))
    bypass_prefix = 'ZonesChW' if branch_suffix == 'ChWBranch' else 'ZonesHW'
    return (
        f'BranchList,{system_name},{bypass_prefix}InletBranch{branches},'
        f'{bypass_prefix}BypassBranch,{bypass_prefix}OutletBranch;\n'
    )


def connector_lines(label: str, branch_suffix: str, floor_count: int) -> list[str]:
    """Create the splitter and mixer objects associated with a hydronic branch family."""
    branches = ''.join(f',Piano{floor}:{branch_suffix}' for floor in range(floor_count))
    prefix = 'ZonesChW' if branch_suffix == 'ChWBranch' else 'ZonesHW'
    return [
        f'Connector:Splitter,{label} Splitter,{prefix}InletBranch{branches},{prefix}BypassBranch;\n',
        f'Connector:Mixer,{label} Mixer,{prefix}OutletBranch{branches},{prefix}BypassBranch;\n',
    ]


def hydronic_branch_lines(floor: int) -> list[str]:
    """Generate the chilled-water and hot-water branch objects for one floor."""
    return [
        f'Branch,Piano{floor}:ChWBranch,,Coil:Cooling:Water,Piano{floor}:FanCoilCoolingCoil,'
        f'Piano{floor}:FanCoilChWInletNode,Piano{floor}:FanCoilChWOutletNode;\n',
        f'Branch,Piano{floor}:HWBranch,,Coil:Heating:Water,Piano{floor}:FanCoilHeatingCoil,'
        f'Piano{floor}:FanCoilHWInletNode,Piano{floor}:FanCoilHWOutletNode;\n',
    ]


def fan_coil_lines(floor: int, schedule_name: str) -> list[str]:
    """Generate the fan-coil and related node objects for one floor."""
    return [
        f'ZoneHVAC:EquipmentConnections,Piano{floor},Piano{floor}:Equipment,Piano{floor}:Inlets,'
        f'Piano{floor}:Exhausts,Piano{floor}:Node,Piano{floor}:OutletNode;\n',
        f'ZoneHVAC:EquipmentList,Piano{floor}:Equipment,SequentialLoad,ZoneHVAC:FourPipeFanCoil,'
        f'Piano{floor}:FanCoil,1,1,,;\n',
        f'NodeList,Piano{floor}:Inlets,Piano{floor}:FanCoilAirOutletNode;\n',
        f'NodeList,Piano{floor}:Exhausts,Piano{floor}:FanCoilAirInletNode;\n',
        f'OutdoorAir:Mixer,Piano{floor}:FanCoilOAMixer,Piano{floor}:FanCoilOAMixerOutletNode,'
        f'Piano{floor}:FanCoilOAInNode,Piano{floor}:FanCoilExhNode,Piano{floor}:FanCoilAirInletNode;\n',
        f'ZoneHVAC:FourPipeFanCoil,Piano{floor}:FanCoil,{schedule_name},ConstantFanVariableFlow,autosize,,,'
        f'autosize,,Piano{floor}:FanCoilAirInletNode,Piano{floor}:FanCoilAirOutletNode,OutdoorAir:Mixer,'
        f'Piano{floor}:FanCoilOAMixer,Fan:ConstantVolume,Piano{floor}:FanCoilFan,Coil:Cooling:Water,'
        f'Piano{floor}:FanCoilCoolingCoil,autosize,0.0,0.001,Coil:Heating:Water,Piano{floor}:FanCoilHeatingCoil,'
        f'autosize,0.0,0.001;\n',
        f'Fan:ConstantVolume,Piano{floor}:FanCoilFan,{schedule_name},0.5,75.0,autosize,0.9,1.0,'
        f'Piano{floor}:FanCoilOAMixerOutletNode,Piano{floor}:FanCoilFanOutletNode;\n',
        f'Coil:Cooling:Water,Piano{floor}:FanCoilCoolingCoil,{schedule_name},autosize,autosize,autosize,autosize,'
        f'autosize,autosize,autosize,Piano{floor}:FanCoilChWInletNode,Piano{floor}:FanCoilChWOutletNode,'
        f'Piano{floor}:FanCoilFanOutletNode,Piano{floor}:FanCoilCCOutletNode,SimpleAnalysis,CrossFlow,,4;\n',
        f'Coil:Heating:Water,Piano{floor}:FanCoilHeatingCoil,{schedule_name},autosize,autosize,'
        f'Piano{floor}:FanCoilHWInletNode,Piano{floor}:FanCoilHWOutletNode,Piano{floor}:FanCoilCCOutletNode,'
        f'Piano{floor}:FanCoilAirOutletNode,UFactorTimesAreaAndDesignWaterFlowRate,autosize,82.2,16.6,71.1,32.2,,11;\n',
    ]


def build_hvac_lines(floor_count: int, schedule_name: str) -> list[str]:
    """Build fan-coil and hydronic loop snippets for all floors."""
    lines = ['\n!- A completamento della Demand side water\n']
    lines.append(branch_list_line('Cooling Demand Side Branches', 'ChWBranch', floor_count))
    lines.extend(connector_lines('Zones ChW', 'ChWBranch', floor_count))
    lines.append(branch_list_line('Heating Demand Side Branches', 'HWBranch', floor_count))
    lines.extend(connector_lines('Zones HW', 'HWBranch', floor_count))
    for floor in range(floor_count):
        lines.extend(hydronic_branch_lines(floor))

    outside_air_nodes = ''.join(f',Piano{floor}:FanCoilOAInNode' for floor in range(floor_count))
    lines.extend(['\n!- A completamento del circuito ARIA\n', f'NodeList,OutsideAirInletNodes{outside_air_nodes};\n'])
    for floor in range(floor_count):
        lines.extend(fan_coil_lines(floor, schedule_name))
    return lines


def run_command(command: list[str], cwd: Optional[Path] = None) -> None:
    """Execute an external command and surface stdout or stderr details when it fails."""
    completed = subprocess.run(
        command,
        cwd=str(cwd) if cwd is not None else None,
        check=False,
        capture_output=True,
        text=True,
    )
    if completed.returncode != 0:
        details = (completed.stderr or completed.stdout or '').strip()
        raise RuntimeError(f'Comando fallito ({completed.returncode}): {" ".join(command)}\n{details}')


def generate_building_lines(
    building_index: int,
    shape,
    template_lines: list[str],
    shape_types: list,
    floors_by_building: list,
    typology_data,
    shapes: list,
    config: SimulationConfig,
) -> tuple[str, list[str]]:
    """Build the full IDF snippet for one building from geometry and typology data."""
    vertices, vertex_count = clean_vertices(shape.points, config)
    adjacent_flags = adjacent_walls(building_index, vertices, shapes, config)
    floor_count = int(floors_by_building[building_index])
    typology = load_typology(typology_data, shape_types[building_index])
    levels = floor_levels(floor_count, typology.floor_height)
    sill_height, window_height, window_ratio_linear = compute_window_parameters(
        vertices=vertices,
        vertex_count=vertex_count,
        adjacent_flags=adjacent_flags,
        typology=typology,
        config=config,
        building_index=building_index,
    )

    lines = (
        build_zone_lines(floor_count)
        + build_envelope_lines(
            vertex_count=vertex_count,
            vertices=vertices,
            floor_count=floor_count,
            floor_levels=levels,
            typology=typology,
            sill_height=sill_height,
            window_height=window_height,
            window_ratio_linear=window_ratio_linear,
            adjacent_flags=adjacent_flags,
            config=config,
        )
        + build_hvac_lines(floor_count, str(typology.schedule_name))
    )
    return f'Edificio{building_index}.idf', template_lines + lines


def run_simulations_for_building(
    idf_path: Path,
    weather_files: list[Path],
    output_dir: Path,
    config: SimulationConfig,
    callback: Optional[Callable[[str], None]] = None,
) -> None:
    """Run EnergyPlus and ReadVarsESO for one building across all selected weather files."""
    energyplus_exe, readvars_exe = energyplus_paths(config)
    building_name = idf_path.stem
    for weather_index, weather_file in enumerate(weather_files):
        weather_output = output_dir / f'{building_name}W{weather_index}'
        weather_output.mkdir(parents=True, exist_ok=True)
        log_message(f'Simula {building_name} con meteo {weather_file.name}', callback)
        run_command([str(energyplus_exe), '-w', str(weather_file), '-d', str(weather_output), str(idf_path)])
        rvi_path = weather_output / 'my.rvi'
        rvi_path.write_text('eplusout.eso\neplusout.csv\n', encoding='utf-8')
        for profile in config.weather_profiles:
            run_command([str(readvars_exe), 'my.rvi', profile], cwd=weather_output)


def execute_simulation(
    config: SimulationConfig,
    callback: Optional[Callable[[str], None]] = None,
    typology_data=None,
) -> Path:
    """Run the full generation + simulation pipeline and return the output directory."""
    pd, shapefile, _, _ = require_runtime_dependencies()
    validate_config(config)
    output_dir = config.output_dir or unique_output_dir(config.base_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    start_time = dt.datetime.now()
    log_message(f'Avvio simulazione alle {start_time.isoformat(timespec="seconds")}', callback)
    log_message(f'Cartella di output: {output_dir}', callback)
    log_message(f'Sorgente tipologie: {typology_source_label(config)}', callback)

    if typology_data is None:
        typology_data = load_typology_data(config, pd)
    reader = shapefile.Reader(str(config.shapefile_path))
    shapes = reader.shapes()
    building_types = field_values(reader, 'edificio-t')
    floors_by_building = field_values(reader, 'num_piani')
    template_lines = config.idf_template_path.read_text(encoding='utf-8').splitlines(keepends=True)

    final_index = len(shapes) if config.end_index is None else min(config.end_index, len(shapes))
    for building_index in range(config.start_index, final_index):
        log_message(f'Genero edificio {building_index + 1}/{len(shapes)}', callback)
        idf_name, idf_lines = generate_building_lines(
            building_index=building_index,
            shape=shapes[building_index],
            template_lines=template_lines,
            shape_types=building_types,
            floors_by_building=floors_by_building,
            typology_data=typology_data,
            shapes=shapes,
            config=config,
        )
        idf_path = output_dir / idf_name
        idf_path.write_text(''.join(idf_lines), encoding='utf-8')
        run_simulations_for_building(idf_path, config.weather_files, output_dir, config, callback)

    if config.generate_reports:
        write_summary_reports(output_dir, config.weather_files, callback)

    elapsed = dt.datetime.now() - start_time
    log_message(f'Completato in {elapsed}', callback)
    return output_dir


def write_optimization_reports(
    output_dir: Path,
    scenario_rows: list[dict[str, float | str]],
    intervention_rows: list[dict[str, str]],
    baseline_summary: dict[str, float],
    pd,
) -> list[Path]:
    """Export detailed scenario, intervention, summary, and Pareto outputs for an optimization run."""
    if not scenario_rows:
        return []
    from openpyxl.chart import Reference, ScatterChart
    from openpyxl.chart.series_factory import SeriesFactory

    results_df = pd.DataFrame(scenario_rows)
    pareto_df = results_df[results_df['pareto'] == 'SI'].copy()
    intervention_df = pd.DataFrame(intervention_rows)
    pareto_points_df = pareto_df[
        [
            'scenario_id',
            'scenario_label',
            'interventi',
            'energy_total_mwh',
            'lca_final_kgco2eq',
            'energy_saved_mwh_vs_baseline',
            'lca_reduction_kgco2eq_vs_baseline',
        ]
    ].sort_values(by='energy_total_mwh')
    summary_df = pd.DataFrame(
        [
            {
                'baseline_energy_mwh': round(baseline_summary['baseline_energy_mwh'], 6),
                'baseline_lca_kgco2eq': round(baseline_summary['baseline_lca_kgco2eq'], 6),
                'scenari_totali': len(results_df),
                'scenari_pareto': int((results_df['pareto'] == 'SI').sum()),
                'miglior_risparmio_energetico_mwh': round(float(results_df['energy_saved_mwh_vs_baseline'].max()), 6),
                'miglior_risparmio_energetico_pct': round(float(results_df['energy_saved_pct_vs_baseline'].max()), 6),
                'miglior_riduzione_lca_kgco2eq': round(float(results_df['lca_reduction_kgco2eq_vs_baseline'].max()), 6),
                'miglior_riduzione_lca_pct': round(float(results_df['lca_reduction_pct_vs_baseline'].max()), 6),
            }
        ]
    )
    pareto_summary_df = pd.DataFrame(pareto_summary_rows(scenario_rows))
    csv_path = output_dir / 'optimization_results.csv'
    interventions_csv_path = output_dir / 'optimization_interventions_by_typology.csv'
    pareto_points_csv_path = output_dir / 'pareto_points.csv'
    workbook_path = output_dir / 'optimization_results.xlsx'
    results_df.to_csv(csv_path, index=False)
    intervention_df.to_csv(interventions_csv_path, index=False)
    pareto_points_df.to_csv(pareto_points_csv_path, index=False)
    with pd.ExcelWriter(workbook_path, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='ScenarioResults', index=False)
        pareto_df.to_excel(writer, sheet_name='ParetoFront', index=False)
        pareto_points_df.to_excel(writer, sheet_name='ParetoPoints', index=False)
        intervention_df.to_excel(writer, sheet_name='InterventiTipologie', index=False)
        pareto_summary_df.to_excel(writer, sheet_name='ParetoSummary', index=False)
        summary_df.to_excel(writer, sheet_name='OptimizationSummary', index=False)

        workbook = writer.book
        pareto_sheet = writer.sheets['ParetoPoints']
        # Export both tabular results and an embedded chart so the Pareto trade-off
        # can be inspected immediately without post-processing in external tools.
        chart = ScatterChart()
        chart.title = 'Frontiera di Pareto'
        chart.x_axis.title = 'Energia totale [MWh]'
        chart.y_axis.title = 'LCA finale [kgCO2eq]'
        chart.style = 2
        chart.height = 10
        chart.width = 18

        scenario_sheet = writer.sheets['ScenarioResults']
        energy_col = results_df.columns.get_loc('energy_total_mwh') + 1
        lca_col = results_df.columns.get_loc('lca_final_kgco2eq') + 1
        all_x = Reference(scenario_sheet, min_col=energy_col, min_row=2, max_row=len(results_df) + 1)
        all_y = Reference(scenario_sheet, min_col=lca_col, min_row=2, max_row=len(results_df) + 1)
        all_series = SeriesFactory(all_y, all_x, title='Tutti gli scenari')
        all_series.marker.symbol = 'circle'
        chart.series.append(all_series)

        if not pareto_points_df.empty:
            pareto_x = Reference(pareto_sheet, min_col=4, min_row=2, max_row=len(pareto_points_df) + 1)
            pareto_y = Reference(pareto_sheet, min_col=5, min_row=2, max_row=len(pareto_points_df) + 1)
            pareto_series = SeriesFactory(pareto_y, pareto_x, title='Pareto')
            pareto_series.graphicalProperties.line.solidFill = 'C00000'
            pareto_series.graphicalProperties.line.width = 28575
            pareto_series.marker.symbol = 'diamond'
            chart.series.append(pareto_series)

        chart_sheet = workbook.create_sheet('ParetoChart')
        chart_sheet.add_chart(chart, 'B2')
    return [csv_path, interventions_csv_path, pareto_points_csv_path, workbook_path]


def execute_optimization(config: SimulationConfig, callback: Optional[Callable[[str], None]] = None) -> Path:
    """Execute the discrete scenario generation, simulation, LCA evaluation, and Pareto filtering workflow."""
    pd, shapefile, _, _ = require_runtime_dependencies()
    validate_config(config)

    base_df = load_template_sheet(config, pd, TEMPLATE_BASE_SHEET)
    intervention_df = load_template_sheet(config, pd, TEMPLATE_INTERVENTION_SHEET)
    variables_df = load_template_sheet(config, pd, TEMPLATE_VARIABLES_SHEET)
    mapping_df = load_template_sheet(config, pd, TEMPLATE_MAPPING_SHEET)
    lca_df = load_template_sheet(config, pd, TEMPLATE_LCA_SHEET)
    parameters_df = load_template_sheet(config, pd, TEMPLATE_PARAMETERS_SHEET)
    parameters = parameter_dict(parameters_df)

    scenarios = build_scenarios(base_df, intervention_df, variables_df)
    if not scenarios:
        raise ValueError('Nessuno scenario di ottimizzazione generato dal template.')

    output_dir = config.output_dir or unique_output_dir(config.base_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    log_message(f'Avvio ottimizzazione nella cartella: {output_dir}', callback)
    log_message(f'Scenari generati: {len(scenarios)}', callback)

    defaults_by_column = variable_defaults_by_column(variables_df)
    lca_factors = lca_factor_table(lca_df)
    reader = shapefile.Reader(str(config.shapefile_path))
    shapes = reader.shapes()
    building_types = field_values(reader, 'edificio-t')
    floors_by_building = field_values(reader, 'num_piani')
    horizon_years = parameter_float(parameters, 'orizzonte_anni', 1.0)

    scenario_rows: list[dict[str, float | str]] = []
    intervention_rows: list[dict[str, str]] = []
    base_by_type = base_df.set_index('Tipologia', drop=False)

    for scenario in scenarios:
        scenario_output_dir = output_dir / scenario.id
        scenario_config = replace(
            config,
            output_dir=scenario_output_dir,
            generate_reports=config.generate_reports,
            optimize=False,
            typology_sheet=TEMPLATE_INTERVENTION_SHEET,
        )
        log_message(f'Elaboro scenario {scenario.label} ({scenario.id})', callback)
        execute_simulation(scenario_config, callback=callback, typology_data=scenario.typology_data)
        intervention_rows.extend(scenario_intervention_rows(scenario, base_df, scenario.typology_data))
        energy_totals = aggregate_energy_results(scenario_output_dir, config.weather_files)
        energy_context = {
            'energia_termica_annua': energy_totals['natural_gas_mwh'] * horizon_years,
            'energia_elettrica_annua': energy_totals['electricity_mwh'] * horizon_years,
            'glass_disposal_thickness_m': parameter_float(parameters, 'glass_disposal_thickness_m', 0.024),
            'opaque_disposal_thickness_m': parameter_float(parameters, 'opaque_disposal_thickness_m', 0.0),
        }

        production_lca = 0.0
        disposal_lca = 0.0
        scenario_indexed = scenario.typology_data.set_index('Tipologia', drop=False)
        for building_index, shape in enumerate(shapes):
            building_type = building_types[building_index]
            floor_count = int(floors_by_building[building_index])
            scenario_typology = load_typology(scenario.typology_data, building_type)
            geometry = building_geometry_summary(
                building_index=building_index,
                shape=shape,
                shapes=shapes,
                floor_count=floor_count,
                typology=scenario_typology,
                config=config,
            )
            base_row = base_by_type.loc[building_type]
            scenario_row = scenario_indexed.loc[building_type]
            for column_name in scenario.selected_columns:
                old_value = base_row[column_name]
                new_value = scenario_row[column_name]
                if old_value == new_value:
                    continue
                category_name = INTERVENTION_COLUMNS[column_name][0]
                for mapping_row in mapping_rows_for(mapping_df, category_name, new_value, 'Production'):
                    production_lca += lca_amount_for_mapping(
                        mapping_row=mapping_row,
                        geometry=geometry,
                        defaults=defaults_by_column,
                        factors=lca_factors,
                        selected_column=column_name,
                        energy_context=energy_context,
                    )
                for mapping_row in mapping_rows_for(mapping_df, category_name, old_value, 'Disposal'):
                    disposal_lca += lca_amount_for_mapping(
                        mapping_row=mapping_row,
                        geometry=geometry,
                        defaults=defaults_by_column,
                        factors=lca_factors,
                        selected_column=column_name,
                        energy_context=energy_context,
                    )

        # Operational impacts are computed at scenario level from the simulated
        # annual energy demand aggregated over the full building cluster.
        use_lca = 0.0
        for mapping_row in mapping_rows_for(mapping_df, 'energia', 'Natural Gas', 'Use'):
            use_lca += lca_amount_for_mapping(
                mapping_row=mapping_row,
                geometry=BuildingGeometrySummary(0.0, 0.0, 0.0, 0.0, 0.0),
                defaults=defaults_by_column,
                factors=lca_factors,
                selected_column='nome della costruzione del tetto',
                energy_context=energy_context,
            )
        for mapping_row in mapping_rows_for(mapping_df, 'energia', 'Electricity', 'Use'):
            use_lca += lca_amount_for_mapping(
                mapping_row=mapping_row,
                geometry=BuildingGeometrySummary(0.0, 0.0, 0.0, 0.0, 0.0),
                defaults=defaults_by_column,
                factors=lca_factors,
                selected_column='nome della costruzione del tetto',
                energy_context=energy_context,
            )

        scenario_rows.append(
            {
                'scenario_id': scenario.id,
                'scenario_label': scenario.label,
                'interventi': ', '.join(INTERVENTION_COLUMNS[column][1] for column in scenario.selected_columns) or 'baseline',
                'energy_total_mwh': round(float(energy_totals['energy_total_mwh']), 6),
                'natural_gas_mwh': round(float(energy_totals['natural_gas_mwh']), 6),
                'electricity_mwh': round(float(energy_totals['electricity_mwh']), 6),
                'lca_production_kgco2eq': round(production_lca, 6),
                'lca_disposal_kgco2eq': round(disposal_lca, 6),
                'lca_use_kgco2eq': round(use_lca, 6),
                'lca_total_kgco2eq': round(production_lca + disposal_lca + use_lca, 6),
                'output_dir': str(scenario_output_dir),
            }
        )

    flags = pareto_flags(scenario_rows)
    for row, flag in zip(scenario_rows, flags):
        row['pareto'] = 'SI' if flag else 'NO'

    scenario_rows, baseline_summary = add_baseline_deltas(scenario_rows)
    created_paths = write_optimization_reports(output_dir, scenario_rows, intervention_rows, baseline_summary, pd)
    if created_paths:
        log_message(
            'Report ottimizzazione generati: ' + ', '.join(path.name for path in created_paths),
            callback,
        )
        log_message(
            'Baseline: '
            f"{baseline_summary['baseline_energy_mwh']:.3f} MWh, "
            f"{baseline_summary['baseline_lca_kgco2eq']:.3f} kgCO2eq",
            callback,
        )
    return output_dir


def parse_args() -> argparse.Namespace:
    """Define and parse the command-line interface of the application."""
    parser = argparse.ArgumentParser(description='Genera modelli IDF da shapefile e lancia EnergyPlus.')
    parser.add_argument('--gui', action='store_true', help='Avvia l\'interfaccia grafica.')
    parser.add_argument('--run-defaults', action='store_true', help='Esegue in CLI usando i percorsi di default.')
    parser.add_argument('--excel', type=Path, help='Percorso del file Excel delle tipologie edilizie.')
    parser.add_argument(
        '--input-template',
        type=Path,
        help='Percorso del template Excel di ottimizzazione da cui leggere le tipologie.',
    )
    parser.add_argument(
        '--typology-sheet',
        default=None,
        help=f'Nome del foglio tipologie da usare nel template (default: {DEFAULT_TEMPLATE_TYPOLOGY_SHEET}).',
    )
    parser.add_argument(
        '--optimize',
        action='store_true',
        help='Esegue l ottimizzazione multi-scenario e calcola la frontiera di Pareto.',
    )
    parser.add_argument('--shapefile', type=Path, help='Percorso dello shapefile (.shp).')
    parser.add_argument('--idf-template', type=Path, help='Percorso del file base IDF.')
    parser.add_argument('--weather', type=Path, action='append', default=None, help='Percorso a un file meteo .epw. Ripetibile.')
    parser.add_argument('--weather-dir', type=Path, help='Cartella da cui leggere automaticamente tutti i file .epw.')
    parser.add_argument('--energyplus-dir', type=Path, help='Cartella di installazione di EnergyPlus.')
    parser.add_argument('--output-dir', type=Path, help='Cartella di output. Se omessa viene creata automaticamente.')
    parser.add_argument('--start-index', type=int, default=DEFAULT_START_INDEX, help='Indice iniziale degli edifici da elaborare.')
    parser.add_argument('--end-index', type=int, help='Indice finale escluso degli edifici da elaborare.')
    parser.add_argument('--skip-reports', action='store_true', help='Non genera i report finali e il DXF unificato.')
    return parser.parse_args()


def config_from_args(args: argparse.Namespace, base_dir: Path) -> SimulationConfig:
    """Build a SimulationConfig instance from parsed command-line arguments."""
    default_config = build_default_config(base_dir)
    weather_files = default_config.weather_files
    if args.weather_dir is not None:
        weather_files = discover_weather_files(args.weather_dir)
    if args.weather:
        weather_files = [path.resolve() for path in args.weather]

    return SimulationConfig(
        excel_path=(args.excel or default_config.excel_path).resolve(),
        shapefile_path=(args.shapefile or default_config.shapefile_path).resolve(),
        idf_template_path=(args.idf_template or default_config.idf_template_path).resolve(),
        weather_files=weather_files,
        energyplus_dir=(args.energyplus_dir or default_config.energyplus_dir).resolve(),
        template_path=(args.input_template or default_config.template_path).resolve()
        if (args.input_template or default_config.template_path) is not None
        else None,
        typology_sheet=args.typology_sheet or default_config.typology_sheet,
        optimize=args.optimize,
        start_index=args.start_index,
        end_index=args.end_index,
        output_dir=args.output_dir.resolve() if args.output_dir else None,
        generate_reports=not args.skip_reports,
        base_dir=base_dir,
    )


class SimulationGui:
    """Minimal Tkinter wrapper around the batch generation workflow."""
    def __init__(self, base_dir: Path) -> None:
        """Initialize the Tkinter window, default values, and background message queue."""
        try:
            import tkinter as tk
            from tkinter import filedialog, messagebox, scrolledtext, ttk
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                'Tkinter non disponibile in questo Python. Usa la CLI oppure installa un Python con supporto Tk.'
            ) from exc

        self._tk = tk
        self._filedialog = filedialog
        self._messagebox = messagebox
        self._ttk = ttk
        self._base_dir = base_dir
        self._default_config = build_default_config(base_dir)
        self._queue: queue.Queue[tuple[str, str]] = queue.Queue()

        self.root = tk.Tk()
        self.root.title('LCADUBS - EnergyPlus GUI')
        self.root.geometry('980x720')

        self.excel_var = tk.StringVar(value=str(self._default_config.excel_path))
        self.template_var = tk.StringVar(value=str(self._default_config.template_path or ''))
        self.typology_sheet_var = tk.StringVar(value=self._default_config.typology_sheet)
        self.shapefile_var = tk.StringVar(value=str(self._default_config.shapefile_path))
        self.idf_var = tk.StringVar(value=str(self._default_config.idf_template_path))
        self.weather_dir_var = tk.StringVar(value=str(base_dir / DEFAULT_WEATHER_DIR))
        self.energyplus_var = tk.StringVar(value=str(self._default_config.energyplus_dir))
        self.output_dir_var = tk.StringVar(value='')
        self.start_var = tk.StringVar(value=str(DEFAULT_START_INDEX))
        self.end_var = tk.StringVar(value='')
        self.status_var = tk.StringVar(value='Pronto')
        self.generate_reports_var = tk.BooleanVar(value=True)
        self.optimize_var = tk.BooleanVar(value=False)

        self._build_layout(scrolledtext)
        self.root.after(150, self._drain_queue)

    def _build_layout(self, scrolledtext_module) -> None:
        """Create all widgets used by the graphical front-end."""
        frame = self._ttk.Frame(self.root, padding=12)
        frame.pack(fill='both', expand=True)
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(12, weight=1)

        self._add_file_row(frame, 0, 'Excel tipologie', self.excel_var, self._browse_excel)
        self._add_file_row(frame, 1, 'Template ottimizzazione', self.template_var, self._browse_template)
        self._ttk.Label(frame, text='Foglio tipologie').grid(row=2, column=0, sticky='w', pady=4)
        self._ttk.Entry(frame, textvariable=self.typology_sheet_var).grid(row=2, column=1, columnspan=2, sticky='ew', pady=4)
        self._ttk.Label(frame, text='Ignorato se il template non e selezionato').grid(
            row=2, column=3, sticky='w', padx=(8, 0), pady=4
        )
        self._add_file_row(frame, 3, 'Shapefile', self.shapefile_var, self._browse_shapefile)
        self._add_file_row(frame, 4, 'Template IDF', self.idf_var, self._browse_idf)
        self._add_file_row(frame, 5, 'Cartella meteo', self.weather_dir_var, self._browse_weather_dir, directory=True)
        self._add_file_row(frame, 6, 'Cartella EnergyPlus', self.energyplus_var, self._browse_energyplus_dir, directory=True)
        self._add_file_row(frame, 7, 'Cartella output', self.output_dir_var, self._browse_output_dir, directory=True)

        self._ttk.Label(frame, text='Start index').grid(row=8, column=0, sticky='w', pady=(8, 4))
        self._ttk.Entry(frame, textvariable=self.start_var, width=12).grid(row=8, column=1, sticky='w', pady=(8, 4))
        self._ttk.Label(frame, text='End index (escluso)').grid(row=8, column=2, sticky='w', padx=(12, 0), pady=(8, 4))
        self._ttk.Entry(frame, textvariable=self.end_var, width=12).grid(row=8, column=3, sticky='w', pady=(8, 4))

        self._ttk.Checkbutton(
            frame,
            text='Genera report finali e DXF unificato',
            variable=self.generate_reports_var,
        ).grid(row=9, column=0, columnspan=4, sticky='w', pady=(6, 4))
        self._ttk.Checkbutton(
            frame,
            text='Esegui ottimizzazione Pareto (baseline + combinazioni interventi)',
            variable=self.optimize_var,
        ).grid(row=10, column=0, columnspan=4, sticky='w', pady=(0, 6))

        buttons = self._ttk.Frame(frame)
        buttons.grid(row=11, column=0, columnspan=4, sticky='w', pady=(10, 10))
        self.run_button = self._ttk.Button(buttons, text='Avvia simulazione', command=self._start_run)
        self.run_button.pack(side='left')
        self._ttk.Button(buttons, text='Chiudi', command=self.root.destroy).pack(side='left', padx=(8, 0))
        self._ttk.Label(buttons, textvariable=self.status_var).pack(side='left', padx=(16, 0))

        self.log_widget = scrolledtext_module.ScrolledText(frame, wrap='word', height=22)
        self.log_widget.grid(row=12, column=0, columnspan=4, sticky='nsew')
        self.log_widget.insert('end', 'Interfaccia pronta. Seleziona i file e premi Avvia simulazione.\n')
        self.log_widget.configure(state='disabled')

    def _add_file_row(self, parent, row: int, label: str, variable, command, directory: bool = False) -> None:
        """Create one labeled row composed of text entry and browse button."""
        self._ttk.Label(parent, text=label).grid(row=row, column=0, sticky='w', pady=4)
        self._ttk.Entry(parent, textvariable=variable).grid(row=row, column=1, columnspan=2, sticky='ew', pady=4)
        button_text = 'Sfoglia cartella' if directory else 'Sfoglia file'
        self._ttk.Button(parent, text=button_text, command=command).grid(row=row, column=3, sticky='ew', padx=(8, 0), pady=4)

    def _browse_excel(self) -> None:
        """Open a file dialog to select the legacy Excel typology workbook."""
        path = self._filedialog.askopenfilename(filetypes=[('Excel', '*.xlsx *.xls')])
        if path:
            self.excel_var.set(path)

    def _browse_template(self) -> None:
        """Open a file dialog to select the optimization template workbook."""
        path = self._filedialog.askopenfilename(filetypes=[('Excel', '*.xlsx *.xls')])
        if path:
            self.template_var.set(path)

    def _browse_shapefile(self) -> None:
        """Open a file dialog to select the building shapefile."""
        path = self._filedialog.askopenfilename(filetypes=[('Shapefile', '*.shp')])
        if path:
            self.shapefile_var.set(path)

    def _browse_idf(self) -> None:
        """Open a file dialog to select the base EnergyPlus IDF template."""
        path = self._filedialog.askopenfilename(filetypes=[('EnergyPlus IDF', '*.idf')])
        if path:
            self.idf_var.set(path)

    def _browse_weather_dir(self) -> None:
        """Open a directory chooser for the folder containing EPW weather files."""
        path = self._filedialog.askdirectory()
        if path:
            self.weather_dir_var.set(path)

    def _browse_energyplus_dir(self) -> None:
        """Open a directory chooser for the EnergyPlus installation folder."""
        path = self._filedialog.askdirectory()
        if path:
            self.energyplus_var.set(path)

    def _browse_output_dir(self) -> None:
        """Open a directory chooser for the desired output directory."""
        path = self._filedialog.askdirectory()
        if path:
            self.output_dir_var.set(path)

    def _append_log(self, message: str) -> None:
        """Append one line to the GUI log widget while preserving read-only behavior."""
        self.log_widget.configure(state='normal')
        self.log_widget.insert('end', message + '\n')
        self.log_widget.see('end')
        self.log_widget.configure(state='disabled')

    def _drain_queue(self) -> None:
        """Consume messages from the worker thread and update the GUI state accordingly."""
        while not self._queue.empty():
            kind, payload = self._queue.get()
            if kind == 'log':
                self._append_log(payload)
            elif kind == 'done':
                self.run_button.configure(state='normal')
                self.status_var.set('Completato')
                self._append_log(payload)
                self._messagebox.showinfo('LCADUBS', payload)
            elif kind == 'error':
                self.run_button.configure(state='normal')
                self.status_var.set('Errore')
                self._append_log(payload)
                self._messagebox.showerror('LCADUBS', payload)
        self.root.after(150, self._drain_queue)

    def _build_config(self) -> SimulationConfig:
        """Collect GUI values and build the runtime configuration object."""
        weather_dir = Path(self.weather_dir_var.get().strip()).expanduser()
        weather_files = discover_weather_files(weather_dir)
        end_value = self.end_var.get().strip()
        template_value = self.template_var.get().strip()
        return SimulationConfig(
            excel_path=Path(self.excel_var.get().strip()).expanduser(),
            shapefile_path=Path(self.shapefile_var.get().strip()).expanduser(),
            idf_template_path=Path(self.idf_var.get().strip()).expanduser(),
            weather_files=weather_files,
            energyplus_dir=Path(self.energyplus_var.get().strip()).expanduser(),
            template_path=Path(template_value).expanduser() if template_value else None,
            typology_sheet=self.typology_sheet_var.get().strip() or DEFAULT_TEMPLATE_TYPOLOGY_SHEET,
            optimize=self.optimize_var.get(),
            start_index=int(self.start_var.get().strip()),
            end_index=int(end_value) if end_value else None,
            output_dir=Path(self.output_dir_var.get().strip()).expanduser() if self.output_dir_var.get().strip() else None,
            generate_reports=self.generate_reports_var.get(),
            base_dir=self._base_dir,
        )

    def _start_run(self) -> None:
        """Validate the current GUI configuration and launch the requested background workflow."""
        try:
            config = self._build_config()
            validate_config(config)
        except Exception as exc:
            self._messagebox.showerror('LCADUBS', str(exc))
            return

        self.run_button.configure(state='disabled')
        self.status_var.set('In esecuzione...')
        self._append_log('Avvio simulazione in background...')

        def worker() -> None:
            """Run the selected workflow in a background thread and push status messages back to the GUI queue."""
            try:
                runner = execute_optimization if config.optimize else execute_simulation
                output_dir = runner(config, callback=lambda msg: self._queue.put(('log', msg)))
                operation = 'Ottimizzazione' if config.optimize else 'Simulazione'
                self._queue.put(('done', f'{operation} completata. Output in: {output_dir}'))
            except Exception as exc:
                self._queue.put(('error', str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def run(self) -> None:
        """Start the Tkinter main loop."""
        self.root.mainloop()


def main() -> None:
    """Choose between GUI and CLI execution modes and launch the requested workflow."""
    base_dir = Path.cwd()
    args = parse_args()
    should_launch_gui = args.gui or not any(
        [
            args.run_defaults,
            args.excel,
            args.input_template,
            args.typology_sheet,
            args.optimize,
            args.shapefile,
            args.idf_template,
            args.weather,
            args.weather_dir,
            args.energyplus_dir,
            args.output_dir,
            args.end_index is not None,
            args.start_index != DEFAULT_START_INDEX,
            args.skip_reports,
        ]
    )
    if should_launch_gui:
        try:
            SimulationGui(base_dir).run()
        except RuntimeError as exc:
            print(f'{LOG_PREFIX} {exc}', file=sys.stderr)
            print(f'{LOG_PREFIX} Esempio CLI: .venv/bin/python main.py --run-defaults', file=sys.stderr)
            raise SystemExit(1) from exc
        return

    config = config_from_args(args, base_dir)
    runner = execute_optimization if config.optimize else execute_simulation
    runner(config)


if __name__ == '__main__':
    main()
